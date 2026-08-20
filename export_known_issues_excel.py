import json
import re
from pathlib import Path

import requests
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

MCP_JSON = Path.home() / ".cursor" / "mcp.json"
OUTPUT_FILE = Path(__file__).resolve().parent / "BIIH_Known_Issues.xlsx"
ACCEPTANCE_CRITERIA_FIELD = "customfield_15048"

TICKET_KEYS = (
    ["BIIH-9", "BIIH-61"]
    + [f"BIIH-{n}" for n in range(148, 173)]
    + ["BIIH-178", "BIIH-179"]
)

COLUMNS = [
    ("Ticket ID", 12),
    ("Title", 45),
    ("Summary", 45),
    ("Description", 70),
    ("Acceptance Criteria", 50),
    ("Type", 14),
    ("Priority", 12),
    ("Labels", 18),
    ("Status", 14),
    ("Created", 20),
    ("Updated", 20),
    ("Reporter", 25),
    ("Assignee", 25),
    ("Jira URL", 55),
    ("Fetch Status", 18),
]


def load_config():
    data = json.loads(MCP_JSON.read_text(encoding="utf-8"))
    env = data["mcpServers"]["mcp-atlassian"]["env"]
    return env["JIRA_URL"].rstrip("/"), env["JIRA_PERSONAL_TOKEN"]


def adf_to_text(node):
    if node is None:
        return ""
    if isinstance(node, str):
        return node
    if isinstance(node, list):
        return "".join(adf_to_text(n) for n in node)
    if not isinstance(node, dict):
        return str(node)

    node_type = node.get("type")
    content = node.get("content", [])

    if node_type == "text":
        return node.get("text", "")
    if node_type == "hardBreak":
        return "\n"
    if node_type == "paragraph":
        return adf_to_text(content) + "\n"
    if node_type == "heading":
        level = node.get("attrs", {}).get("level", 1)
        return ("#" * level) + " " + adf_to_text(content).strip() + "\n"
    if node_type == "bulletList":
        lines = []
        for item in content:
            if item.get("type") == "listItem":
                text = adf_to_text(item.get("content", [])).strip()
                for line in text.splitlines():
                    lines.append(f"- {line}" if line else "-")
        return "\n".join(lines) + ("\n" if lines else "")
    if node_type == "orderedList":
        lines = []
        start = node.get("attrs", {}).get("order", 1)
        idx = start
        for item in content:
            if item.get("type") == "listItem":
                text = adf_to_text(item.get("content", [])).strip()
                for line in text.splitlines():
                    lines.append(f"{idx}. {line}" if line else f"{idx}.")
                    idx += 1
        return "\n".join(lines) + ("\n" if lines else "")
    if node_type in ("doc", "listItem", "table", "tableRow", "tableCell", "mediaSingle"):
        return adf_to_text(content)
    return adf_to_text(content)


def field_to_text(value):
    if value is None:
        return ""
    if isinstance(value, str):
        return value
    if isinstance(value, dict):
        if value.get("type") == "doc":
            return adf_to_text(value).strip()
        if "content" in value and isinstance(value["content"], list):
            return adf_to_text(value).strip()
        if "name" in value:
            return value.get("name", "")
        if "displayName" in value:
            return value.get("displayName", "")
        return json.dumps(value, ensure_ascii=False)
    return str(value)


def extract_acceptance_criteria_from_description(description_text):
    if not description_text:
        return "Not Found"
    pattern = (
        r"(?ims)(?:^|\n)\s*(?:#{1,6}\s*)?"
        r"(Acceptance Criteria|AC|Requirements|Success Criteria)\s*:?\s*\n"
        r"(.*?)(?=\n\s*(?:#{1,6}\s*)?[A-Z][A-Za-z /_-]{2,}\s*:?\s*\n|\Z)"
    )
    match = re.search(pattern, description_text)
    return match.group(2).strip() if match else "Not Found"


def fetch_issue(base_url, token, issue_key):
    url = f"{base_url}/rest/api/2/issue/{issue_key}"
    params = {"expand": "renderedFields"}
    headers = {"Authorization": f"Bearer {token}", "Accept": "application/json"}
    response = requests.get(url, headers=headers, params=params, timeout=60)
    response.raise_for_status()
    return response.json()


def parse_issue(base_url, issue_key, issue):
    fields = issue.get("fields", {})

    summary = field_to_text(fields.get("summary")) or "Not Found"
    description = field_to_text(fields.get("description")) or "Not Found"
    rendered = issue.get("renderedFields", {})
    if description == "Not Found" and rendered.get("description"):
        description = re.sub(r"<[^>]+>", "", rendered["description"]).strip() or "Not Found"

    acceptance = field_to_text(fields.get(ACCEPTANCE_CRITERIA_FIELD))
    if not acceptance:
        acceptance = extract_acceptance_criteria_from_description(description)
    if not acceptance:
        acceptance = "Not Found"

    labels = fields.get("labels") or []
    labels_text = ", ".join(labels) if labels else "None"

    return {
        "Ticket ID": issue_key,
        "Title": summary,
        "Summary": summary,
        "Description": description,
        "Acceptance Criteria": acceptance,
        "Type": (fields.get("issuetype") or {}).get("name") or "Not Found",
        "Priority": (fields.get("priority") or {}).get("name") or "Not Found",
        "Labels": labels_text,
        "Status": (fields.get("status") or {}).get("name") or "Not Found",
        "Created": field_to_text(fields.get("created")) or "Not Found",
        "Updated": field_to_text(fields.get("updated")) or "Not Found",
        "Reporter": field_to_text(fields.get("reporter")) or "Not Found",
        "Assignee": field_to_text(fields.get("assignee")) or "Unassigned",
        "Jira URL": f"{base_url}/browse/{issue_key}",
        "Fetch Status": "Fetched",
    }


def build_workbook(rows):
    wb = Workbook()
    ws = wb.active
    ws.title = "Known Issues"

    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for col_idx, (header, width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}{max(len(rows) + 1, 2)}"

    wrap = Alignment(vertical="top", wrap_text=True)
    for row_idx, row in enumerate(rows, start=2):
        for col_idx, (header, _) in enumerate(COLUMNS, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=row.get(header, ""))
            cell.alignment = wrap

    ws.row_dimensions[1].height = 24
    return wb


def main():
    base_url, token = load_config()
    rows = []

    for issue_key in TICKET_KEYS:
        try:
            issue = fetch_issue(base_url, token, issue_key)
            rows.append(parse_issue(base_url, issue_key, issue))
            print(f"Fetched {issue_key}")
        except requests.HTTPError as exc:
            status = exc.response.status_code if exc.response is not None else "unknown"
            rows.append(
                {
                    "Ticket ID": issue_key,
                    "Title": "Not Found",
                    "Summary": "Not Found",
                    "Description": "Not Found",
                    "Acceptance Criteria": "Not Found",
                    "Type": "Not Found",
                    "Priority": "Not Found",
                    "Labels": "Not Found",
                    "Status": "Not Found",
                    "Created": "Not Found",
                    "Updated": "Not Found",
                    "Reporter": "Not Found",
                    "Assignee": "Not Found",
                    "Jira URL": f"{base_url}/browse/{issue_key}",
                    "Fetch Status": f"Inaccessible (HTTP {status})",
                }
            )
            print(f"Missing {issue_key}: HTTP {status}")
        except requests.RequestException as exc:
            rows.append(
                {
                    "Ticket ID": issue_key,
                    "Title": "Not Found",
                    "Summary": "Not Found",
                    "Description": "Not Found",
                    "Acceptance Criteria": "Not Found",
                    "Type": "Not Found",
                    "Priority": "Not Found",
                    "Labels": "Not Found",
                    "Status": "Not Found",
                    "Created": "Not Found",
                    "Updated": "Not Found",
                    "Reporter": "Not Found",
                    "Assignee": "Not Found",
                    "Jira URL": f"{base_url}/browse/{issue_key}",
                    "Fetch Status": f"Error: {exc}",
                }
            )
            print(f"Error {issue_key}: {exc}")

    wb = build_workbook(rows)
    wb.save(OUTPUT_FILE)
    print(str(OUTPUT_FILE))
    print(f"Total rows: {len(rows)}")


if __name__ == "__main__":
    main()
